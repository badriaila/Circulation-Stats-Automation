# Excel 
from openpyxl import Workbook 
from openpyxl import load_workbook
from datetime import date

todays_date = date.today()
print(todays_date)


#year
year = todays_date.year

#month

month = todays_date.month
prev_month = month-1

months = [["Jan","Feb","Mar","Apr","May","June","july","Aug","Sept","Oct","Nov","Dec"],["January","February","March","April","May","June","July","August","September","October","November","December"]]

prev_month_short = months[0][prev_month-1]
prev_month_full = months[1][prev_month-1]



#loading Workbook
wb = load_workbook(filename="Monthly Equipment Checkout_"+str(prev_month_short)+"_"+str(todays_date.year)+".xlsx")

#selecting sheet
sheet = wb["Monthly Equipment Checkout"]

#Getting the Title brief data and appending to all_data
i=2
all_data=[]
for row in sheet.iter_rows():
        all_data.append(sheet["D"+str(i)].value)
        i=i+1

#Calculating fiscal year Works until 2099 if you would like to increase it further in the year 2100 change the value 2000 below to 2100
if prev_month >=6:
        this_year = str(todays_date.year - 2000)
        next_year = str(todays_date.year - 2000 +1)
else:
        this_year = str(todays_date.year - 2000 -1)
        next_year = str(todays_date.year - 2000)


fiscal_year = this_year+"/"+next_year
print(fiscal_year)
fiscal_year_full = "FY "+fiscal_year



#Count
laptop_count = all_data.count("LAPTOP ")
headset_count = all_data.count("HEADSET ")
cable_count= all_data.count("CABLE ")
usbDrive_count = all_data.count("USB DRIVE ")
mouse_count = all_data.count("MOUSE ")
calculator_count = all_data.count("CALCULATOR ")
snowballUsbMicrophone_count = all_data.count("Snowball USB Microphone ")
graphicCalculator_count = all_data.count("GRAPHING CALCULATOR ")
lohaSmartPhoneMountTripod_count = all_data.count("LOHA Smartphone Mount/Tripod ")
boseHeadphones_count = all_data.count("BOSE HEADPHONES ")
ipad_count = all_data.count("IPAD ")
goProHero3Camera_count = all_data.count("GoPro Hero 3 Camera ")
canonVixiaHFM500HDCamcorder_count = all_data.count("Canon Vixia HF M500 HD Camcorder ")
sonyHandycamHDRCX455_count = all_data.count("Sony Handycam HDR-CX455 ")

#A list of lists
count_ = [[fiscal_year_full,year,prev_month_full,laptop_count, headset_count,cable_count, usbDrive_count, mouse_count,calculator_count,snowballUsbMicrophone_count,graphicCalculator_count,lohaSmartPhoneMountTripod_count, boseHeadphones_count,ipad_count,goProHero3Camera_count,canonVixiaHFM500HDCamcorder_count,sonyHandycamHDRCX455_count]]



# Google Sheets
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account


# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = 'c-space-project-key.json'
creds = None
creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)


# The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '1xFEX2JNo9C4W8stz4K68riT6HZZdu2OX-KSyxkhPcZw'
service = build('sheets', 'v4', credentials=creds)

# Call the Sheets API
sheet = service.spreadsheets()



res = sheet.get(spreadsheetId=SAMPLE_SPREADSHEET_ID, fields='sheets(data/rowData/values/userEnteredValue,properties(index,sheetId,title))').execute()


#Circulation Checkout Sheet
sheetIndex = 9
sheetName = res['sheets'][sheetIndex]['properties']['title']
print(sheetName)
lastRow = len(res['sheets'][sheetIndex]['data'][0]['rowData'])
lastColumn = max([len(e['values']) for e in res['sheets'][sheetIndex]['data'][0]['rowData'] if e])
print(lastColumn)
row = "A"+str(lastRow+1)
print(row)
# request = sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID, range="Circulation Checkout!"+row, insertDataOption="INSERT_ROWS", valueInputOption="RAW", body={"values":count_})
# response = request.execute()

# print(response)

#Circulation Checkout update in overall stats sheet
sheetName = 'Overall Stats FY' + fiscal_year
print(sheetName)
lastRow = len(res['sheets'][sheetIndex]['data'][0]['rowData'])
lastColumn = max([len(e['values']) for e in res['sheets'][sheetIndex]['data'][0]['rowData'] if e])
print(lastColumn)
row = "A"+str(lastRow+1)
print(row)





